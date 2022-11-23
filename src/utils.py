import os
import openpyxl

def get_files(path):
    xlsx_files = list() 
    csv_files = list()
    for (root, _, files) in os.walk(path):
        for filename in files:
            if filename.endswith('xlsx'):
                xlsx_files.append("{}\{}".format(str(root), filename))
            elif filename.endswith('csv'):
                csv_files.append(filename)
    return xlsx_files, csv_files

def get_files_with_same_headers(xlsx_files):
    columns_headers_sets = {}
    for fileAddress in xlsx_files:
        headers_list, _, _ = get_headers_and_sheet(fileAddress)
        name_of_key = str(headers_list)
        if name_of_key not in columns_headers_sets:
            columns_headers_sets[name_of_key] = list()
        columns_headers_sets[name_of_key].append(fileAddress)
    return list(columns_headers_sets.values())

def get_headers_and_sheet(fileAdress):
    wb_obj = openpyxl.load_workbook(fileAdress, read_only=True)
    sheet_obj = wb_obj.active
    headers_list = list()
    headers_row = 1
    for row in range(1, sheet_obj.max_row):
        temp_headers_list = list()
        for column in range(1, sheet_obj.max_column+1):
            temp_headers_list.append(sheet_obj.cell(row = row, column = column).value)
        if all(type(header) == str for header in temp_headers_list[0:4]):
            headers_list = temp_headers_list.copy()
            headers_row = row
            break
    return headers_list, headers_row, sheet_obj
