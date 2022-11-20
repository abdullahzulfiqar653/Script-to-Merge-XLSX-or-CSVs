import os
import xlrd
import openpyxl
import pandas as pd


list_of_files = os.listdir('../samples')
xlsx_files = list()
csv_files = list()
for filename in list_of_files:
    if "xlsx" in filename:
        xlsx_files.append(filename)
    if 'csv' in filename:
        csv_files.append(filename)


for filename in xlsx_files:
    address = "../samples/{}".format(filename)
    wb_obj = openpyxl.load_workbook(address)
    sheet_obj = wb_obj.active



for i in range(sheet_obj.max_column):
    print(sheet_obj.cell(row = 5, column = i+1).value)
# print("filename: ", xlsx_files)
# print("csv files:", csv_files)
